from pathlib import Path
import pandas as pd
import requests as rq 
from localidades import nacional, estadual, municipal
import ssl
from Google import Create_Service
import openpyxl
from openpyxl.styles import Font, Border, Side
from ajustar_planilha import ajustar_colunas, ajustar_bordas
from Drive import add_arquivos_a_pasta

api_estadual = f'https://servicodados.ibge.gov.br/api/v3/agregados/5457/periodos/2014|2015|2016|2017|2018|2019|2020|2021|2022|2023/variaveis/8331|216|214|112?{estadual}&classificacao=782[40092,45982,40099,40101,40102,40136,40104,40137,40138,40139,40106,40143,40145,40112,40114,40149,40150,40151,40152,40261,40118,40119,40262,40263,40120,40121,40122,40266,40269,40124,40125,40271,40126,40127,40273,40274]'
ROOT_PATH = Path(__file__).parent

lista_cod_prod = [40092,45982,40099,40101,40102,40136,40104,40137,40138,40139,40106,40143,40145,40112,40114,40149,40150,40151,40152,40261,40118,40119,40262,40263,40120,40121,40122,40266,40269,40124,40125,40271,40126,40127,40273,40274]
class TLSAdapter(rq.adapters.HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.set_ciphers("DEFAULT@SECLEVEL=1")
        ctx.options |= 0x4   # OP_LEGACY_SERVER_CONNECT
        kwargs["ssl_context"] = ctx
        return super(TLSAdapter, self).init_poolmanager(*args, **kwargs)

def requisitando_dados(api):
    with rq.session() as s:
        s.mount("https://", TLSAdapter())
        dados_brutos_api = s.get(api, verify=True)
    
    # Verificação se a solicitação foi bem-sucedida antes de continuar
    if dados_brutos_api.status_code != 200:
        raise Exception(f"A solicitação à API falhou com o código de status: {dados_brutos_api.status_code}")

    # Verificação se a resposta pode ser convertida para JSON
    try:
        dados_brutos = dados_brutos_api.json()
    except Exception as e:
        raise Exception(f"Erro ao analisar a resposta JSON da API: {str(e)}")

    # Verificação se a resposta contém os dados esperados
    if len(dados_brutos) < 4:
        raise Exception("A resposta da API não contém dados suficientes.")
    
    if dados_brutos_api.status_code == 500:
        raise Exception(f"Os dados passou de 100.0000 por isso o codigo de: {dados_brutos_api.status_code}")

    dados_brutos_8331 = dados_brutos[0]
    dados_brutos_216 = dados_brutos[1]
    dados_brutos_214 = dados_brutos[2]
    dados_brutos_112 = dados_brutos[3]
    return dados_brutos_8331, dados_brutos_216, dados_brutos_214, dados_brutos_112

def tratando_dados(dados_brutos_8331, dados_brutos_216, dados_brutos_214, dados_brutos_112):
    dados_limpos_8331 = []
    dados_limpos_216 = []
    dados_limpos_214 = []
    dados_limpos_112 = []

    variaveis = [dados_brutos_8331, dados_brutos_216, dados_brutos_214, dados_brutos_112]

    for i in variaveis:
        id_tabela = i['id']
        variavel = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
        
        if unidade == '':
            unidade = 'Hectares'

        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']
            
            for iii in dados_produto:
                dados_id_produto = iii['categoria']

                for id_produto, nome_produto in dados_id_produto.items():
                    nome_produto = nome_produto.replace('Abacaxi*', 'Abacaxi').replace(' (em caroço)', '').replace(' (em casca)', '').replace(' (em grão) Total', '')\
                    .replace('Coco-da-baía*', 'Coco-da-baía').replace(' (em grão)', '').replace(' (cacho)', '').replace(' (em amêndoa)', '').replace(' (fruto seco)', '')\
                    .replace('Borracha (látex coagulado)', 'Borracha látex coagulado').replace('Borracha (látex líquido)', 'Borracha látex líquido').replace(' (folha verde)', '')\
                    .replace(' (cacho de coco)', '').replace(' (em folha)', '').replace(' (semente)', '').replace(' (fibra)', '').replace(' (baga)', '')
                    
                    for iv in dados_producao:
                        id = iv['localidade']['id']
                        nome = iv['localidade']['nome'].replace(' (MT)', '')
                        dados_ano_producao = iv['serie'] 
                        
                        for ano, producao in dados_ano_producao.items():
                            producao = producao.replace('-', '0').replace('...', '0')
                            
                            dict = {

                                'id': id,
                                'nome': nome,
                                'id_produto': id_produto,
                                'produto': nome_produto,
                                variavel: producao,
                                'ano': f'01/01/{ano}'   
                            }
                           
                            if id_tabela == '8331':
                                dados_limpos_8331.append(dict)
                            elif id_tabela == '216':
                                dados_limpos_216.append(dict)
                            elif id_tabela == '214':
                                dados_limpos_214.append(dict)
                            elif id_tabela == '112':
                                dados_limpos_112.append(dict)


    return dados_limpos_8331, dados_limpos_216, dados_limpos_214, dados_limpos_112

def executando_municipal():
    lista_dados_8331 = [] 
    lista_dados_216 = []
    lista_dados_214 = []
    lista_dados_112 = []
    for codigo in lista_cod_prod:
        api_municipal = f'https://servicodados.ibge.gov.br/api/v3/agregados/5457/periodos/2013|2014|2015|2016|2017|2018|2019|2020|2021|2022|2023/variaveis/8331|216|214|112?localidades=N6[5100102,5100201,5100250,5100300,5100359,5100409,5100508,5100607,5100805,5101001,5101209,5101258,5101308,5101407,5101605,5101704,5101803,5101852,5101902,5102504,5102603,5102637,5102678,5102686,5102694,5102702,5102793,5102850,5103007,5103056,5103106,5103205,5103254,5103304,5103353,5103361,5103379,5103403,5103437,5103452,5103502,5103601,5103700,5103809,5103858,5103908,5103957,5104104,5104203,5104500,5104526,5104542,5104559,5104609,5104807,5104906,5105002,5105101,5105150,5105176,5105200,5105234,5105259,5105309,5105507,5105580,5105606,5105622,5105903,5106000,5106109,5106158,5106174,5106182,5106190,5106208,5106216,5106224,5106232,5106240,5106257,5106265,5106273,5106281,5106299,5106307,5106315,5106372,5106422,5106455,5106505,5106653,5106703,5106752,5106778,5106802,5106828,5106851,5107008,5107040,5107065,5107107,5107156,5107180,5107198,5107206,5107248,5107263,5107297,5107305,5107354,5107404,5107578,5107602,5107701,5107743,5107750,5107768,5107776,5107792,5107800,5107859,5107875,5107883,5107909,5107925,5107941,5107958,5108006,5108055,5108105,5108204,5108303,5108352,5108402,5108501,5108600,5108808,5108857,5108907,5108956]&classificacao=782[{codigo}]'       
        variavel8331municipal, variavel_216municipal, variavel214municipal, variavel112municipal = requisitando_dados(api_municipal)
        if len(variavel8331municipal) == 0 and len(variavel_216municipal) == 0 and len(variavel214municipal) == 0 and len(variavel112municipal) == 0:
            lista_dados_8331, lista_dados_216, lista_dados_214, lista_dados_112 = tratando_dados(variavel8331municipal, variavel_216municipal,variavel214municipal, variavel112municipal)

        else:
            novos_dados_8331, novos_dados_216, novos_dados_214, novos_dados_112 = tratando_dados(variavel8331municipal, variavel_216municipal,variavel214municipal, variavel112municipal)
            lista_dados_8331.extend(novos_dados_8331)
            lista_dados_216.extend(novos_dados_216)
            lista_dados_214.extend(novos_dados_214)
            lista_dados_112.extend(novos_dados_112)
    
    return  lista_dados_8331,lista_dados_216,lista_dados_214, lista_dados_112
def gerando_dataframe(dados_limpos_8331, dados_limpos_216, dados_limpos_214, dados_limpos_112):

    df8331 = pd.DataFrame(dados_limpos_8331)
    df216 = pd.DataFrame(dados_limpos_216)
    df214 = pd.DataFrame(dados_limpos_214)
    df112 = pd.DataFrame(dados_limpos_112)

    dataframe = pd.merge(df8331, df216, on=['id', 'nome', 'id_produto', 'produto', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df214, on=['id', 'nome', 'id_produto', 'produto', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df112, on=['id', 'nome', 'id_produto', 'produto', 'ano'], how='inner')
    dataframe['Área plantada ou destinada à colheita'] = dataframe['Área plantada ou destinada à colheita'].astype(float)
    dataframe['Área colhida'] = dataframe['Área colhida'].astype(float)
    dataframe['Quantidade produzida'] = dataframe['Quantidade produzida'].astype(float)
    return dataframe

def coluna_cultura(dataframe):
    
    culturas_temporarias = ['Abacaxi', 'Algodão herbáceo', 'Amendoim', 'Arroz', 'Batata-doce',
                            'Cana-de-açúcar', 'Feijão', 'Girassol', 'Mamona', 'Mandioca',
                            'Melancia', 'Melão', 'Milho', 'Soja', 'Sorgo', 'Tomate', 'Trigo']
    
    culturas_permanentes = ['Açaí', 'Algodão arbóreo', 'Banana', 'Borracha látex coagulado', 'Borracha látex líquido', 'Cacau', 'Café', 'Castanha de caju',
                            'Coco-da-baía', 'Goiaba', 'Guaraná', 'Laranja', 'Limão', 'Mamão', 'Manga',
                            'Maracujá', 'Palmito', 'Pimenta-do-reino', 'Tangerina', 'Urucum', 'Uva']
    
    
    for index, row in dataframe.iterrows():
        if row['produto'] in culturas_temporarias:
            dataframe.at[index, 'cultura'] = 'Temporária'
        elif row['produto'] in culturas_permanentes:
            dataframe.at[index, 'cultura'] = 'Permanente'
    
    return dataframe


#PARTE ESTADUAL
dados_brutos_8331, dados_brutos_216, dados_brutos_214, dados_brutos_112 = requisitando_dados(api_estadual)
dados_limpos_8331, dados_limpos_216, dados_limpos_214, dados_limpos_112 = tratando_dados(dados_brutos_8331, dados_brutos_216, dados_brutos_214, dados_brutos_112)
dataframe = gerando_dataframe(dados_limpos_8331, dados_limpos_216, dados_limpos_214, dados_limpos_112)
df5457_estadual = coluna_cultura(dataframe)
df5457_estadual.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM_5457_ESTADUAL.xlsx', index=False)
#print(df5457_estadual)
dados_limpos_8331_muni, dados_limpos_216_muni, dados_limpos_214_muni, dados_limpos_112_muni = executando_municipal()
dataframe_muni = gerando_dataframe(dados_limpos_8331_muni, dados_limpos_216_muni, dados_limpos_214_muni, dados_limpos_112_muni)
df5457_municipal = coluna_cultura(dataframe_muni)
df5457_municipal.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM_5457_MUNICIPAL.xlsx', index=False)


# CARREGA A PLANILHA DO PAM 5457 E FAZ AS ALTERAÇÕES ESTRUTURAIS DA PLANILHA

wb_5457_estadual = openpyxl.load_workbook("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM_5457_ESTADUAL.xlsx")  

ws_5457_estadual = wb_5457_estadual.active

lista_ws = [ws_5457_estadual]
lista_wb = [wb_5457_estadual]
for ws, wb in zip(lista_ws, lista_wb):
    ajustar_colunas(ws)
    ajustar_bordas(wb)

wb_5457_estadual.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM_5457_ESTADUAL.xlsx")

#TESTANDO JUNTAR TODAS PLANILHAS EM UMA SÓ
planilha_principal = openpyxl.Workbook()
wb_5457_estadual = openpyxl.load_workbook("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM_5457_ESTADUAL.xlsx")  
wb_5457_municipal = openpyxl.load_workbook("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM_5457_MUNICIPAL.xlsx")  
aba_5457_estadual = planilha_principal.create_sheet("PAM 5457 ESTADUAL")
aba_5457_municipal = planilha_principal.create_sheet("PAM 5457 MUNICIPAL")


# Copiar os dados da segunda planilha para a nova planilha
for linha in wb_5457_estadual.active.iter_rows(values_only=True):
    aba_5457_estadual.append(linha)
    
for linha in wb_5457_municipal.active.iter_rows(values_only=True):
    aba_5457_municipal.append(linha)

for aba in planilha_principal.sheetnames:
    if aba not in ["PAM 5457 ESTADUAL", "PAM 5457 MUNICIPAL"]:
        del planilha_principal[aba]


lista_abas = [aba_5457_estadual, aba_5457_municipal]
for abas in lista_abas:
    ajustar_colunas(abas)


planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM.xlsx")
worksheet = planilha_principal.active
df = pd.read_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM.xlsx')

ajustar_bordas(planilha_principal)

planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\PAM.xlsx")
print("ESTOU EXECUTANDO SQL.PY")
'''
#Faz autenticação do google drive para jogar os arquivos gerados no codigo python
CLIENT_SECRET_FILE = 'credencials.json'
API_NAME = 'drive'
API_VERSION = 'v3'
SCOPES = ["https://www.googleapis.com/auth/drive"]

service = Create_Service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)

#PASSA O PAM PARA O DRIVE
file_id = "1KSR_XbhD-SEL9Wpu-f8SEZT0ETGhGeLb"
FILE_NAMES = ["PAM_5457_NACIONAL.xlsx", "PAM_5457_ESTADUAL.xlsx", "PAM.xlsx"]
MIME_TYPES = ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]

add_arquivos_a_pasta(FILE_NAMES, MIME_TYPES, service, file_id)
'''
#ABRE O ARQUIVO SQL.PY E EXECUTA TODOS OS COMANDOS DENTRO DELE
if __name__ == '__main__':
    from sql import executar_sql 
    executar_sql()

